from __future__ import annotations

from pathlib import Path
from typing import Iterable

import json
import zipfile
from xml.sax.saxutils import escape

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = ROOT / "outputs" / "field-mapping"
OUTPUT_PATH = OUTPUT_DIR / "KTNVL_K2_Field_Mapping_Review.xlsx"
JSON_PATH = OUTPUT_DIR / "KTNVL_K2_Field_Mapping_Review.json"


if Side is not None:
    thin = Side(style="thin", color="D9D4C7")
    header_fill = PatternFill("solid", fgColor="1F7A6C")
    section_fill = PatternFill("solid", fgColor="E8F6F2")
    header_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
else:
    thin = header_fill = section_fill = header_font = bold_font = None


FIELD_HEADERS = [
    "STT",
    "Module",
    "Màn hình",
    "Nhóm field",
    "Tên field hiển thị",
    "Field code",
    "Mô tả nghiệp vụ",
    "Kiểu dữ liệu",
    "Bắt buộc?",
    "Loại nhập liệu",
    "Giữ/Bỏ",
    "Nguồn dữ liệu",
    "Dropdown list code",
    "Công thức / Logic",
    "Phụ thuộc",
    "Ghi chú user",
]


def autosize(ws):
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = min(max(widths.get(cell.column, 0), len(value) + 2), 42)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def style_header(ws, row=1):
    for cell in ws[row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def col_letter(index: int) -> str:
    result = ""
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def sheet_xml(headers: list[str], rows: list[list[str]], subtitle: str) -> str:
    all_rows = [[subtitle], headers] + rows
    xml_rows = []
    max_cols = max(len(headers), 1)

    for row_idx, row in enumerate(all_rows, start=1):
        cells = []
        for col_idx, value in enumerate(row, start=1):
            cell_ref = f"{col_letter(col_idx)}{row_idx}"
            style = ""
            if row_idx == 1:
                style = ' s="1"'
            elif row_idx == 2:
                style = ' s="2"'
            safe_value = escape("" if value is None else str(value))
            cells.append(
                f'<c r="{cell_ref}" t="inlineStr"{style}><is><t xml:space="preserve">{safe_value}</t></is></c>'
            )
        xml_rows.append(f'<row r="{row_idx}">{"".join(cells)}</row>')

    dimension = f"A1:{col_letter(max_cols)}{len(all_rows)}"
    merge_cell = f"A1:{col_letter(max_cols)}1"

    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <dimension ref="{dimension}"/>
  <sheetViews>
    <sheetView workbookViewId="0">
      <pane ySplit="2" topLeftCell="A3" activePane="bottomLeft" state="frozen"/>
      <selection pane="bottomLeft"/>
    </sheetView>
  </sheetViews>
  <sheetFormatPr defaultRowHeight="18"/>
  <cols>
    <col min="1" max="{max_cols}" width="20" customWidth="1"/>
  </cols>
  <sheetData>
    {"".join(xml_rows)}
  </sheetData>
  <mergeCells count="1">
    <mergeCell ref="{merge_cell}"/>
  </mergeCells>
</worksheet>"""


def workbook_xml(sheet_names: list[str]) -> str:
    sheets = []
    for idx, name in enumerate(sheet_names, start=1):
        sheets.append(f'<sheet name="{escape(name)}" sheetId="{idx}" r:id="rId{idx}"/>')
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets>
    {"".join(sheets)}
  </sheets>
</workbook>"""


def workbook_rels_xml(sheet_count: int) -> str:
    rels = []
    for idx in range(1, sheet_count + 1):
        rels.append(
            f'<Relationship Id="rId{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx}.xml"/>'
        )
    rels.append(
        f'<Relationship Id="rId{sheet_count + 1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>'
    )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  {"".join(rels)}
</Relationships>"""


def root_rels_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""


def content_types_xml(sheet_count: int) -> str:
    overrides = [
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
        '<Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
    ]
    for idx in range(1, sheet_count + 1):
        overrides.append(
            f'<Override PartName="/xl/worksheets/sheet{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        )
    return f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  {"".join(overrides)}
</Types>"""


def styles_xml() -> str:
    return """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="3">
    <font><sz val="11"/><name val="Calibri"/></font>
    <font><b/><sz val="11"/><name val="Calibri"/></font>
    <font><b/><color rgb="FFFFFFFF"/><sz val="11"/><name val="Calibri"/></font>
  </fonts>
  <fills count="4">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFE8F6F2"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF1F7A6C"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="1">
    <border><left/><right/><top/><bottom/><diagonal/></border>
  </borders>
  <cellStyleXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
  </cellStyleXfs>
  <cellXfs count="3">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment vertical="top" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="0" xfId="0" applyAlignment="1" applyFill="1" applyFont="1"><alignment vertical="center" wrapText="1"/></xf>
    <xf numFmtId="0" fontId="2" fillId="3" borderId="0" xfId="0" applyAlignment="1" applyFill="1" applyFont="1"><alignment horizontal="center" vertical="center" wrapText="1"/></xf>
  </cellXfs>
  <cellStyles count="1">
    <cellStyle name="Normal" xfId="0" builtinId="0"/>
  </cellStyles>
</styleSheet>"""


def build_xlsx_from_payload(payload: dict) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    sheets = payload["sheets"]

    with zipfile.ZipFile(OUTPUT_PATH, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types_xml(len(sheets)))
        zf.writestr("_rels/.rels", root_rels_xml())
        zf.writestr("xl/workbook.xml", workbook_xml([sheet["name"] for sheet in sheets]))
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml(len(sheets)))
        zf.writestr("xl/styles.xml", styles_xml())

        for idx, sheet in enumerate(sheets, start=1):
            rows: list[list[str]] = []
            for row in sheet["rows"]:
                if isinstance(row, dict):
                    if sheet["name"] in {"01_LSX_Fields", "02_NKNVL_Fields", "03_TonHopTho", "04_HaoHut", "05_GiaDinhMuc"}:
                        rows.append([
                            str(len(rows) + 1),
                            str(row.get("module", "")),
                            str(row.get("screen", "")),
                            str(row.get("group", "")),
                            str(row.get("label", "")),
                            str(row.get("code", "")),
                            str(row.get("description", "")),
                            str(row.get("data_type", "")),
                            str(row.get("required", "")),
                            str(row.get("input_type", "")),
                            str(row.get("decision", "Chua xac nhan")),
                            str(row.get("source", "")),
                            str(row.get("dropdown_code", "")),
                            str(row.get("formula", "")),
                            str(row.get("depends_on", "")),
                            str(row.get("notes", "")),
                        ])
                    elif sheet["name"] == "06_DropdownLists":
                        rows.append([
                            str(len(rows) + 1),
                            str(row.get("group", "")),
                            str(row.get("module", "")),
                            str(row.get("field_code", "")),
                            str(row.get("value", "")),
                            str(row.get("label", "")),
                            str(row.get("note", "")),
                        ])
                else:
                    rows.append([str(value) for value in row])

            zf.writestr(
                f"xl/worksheets/sheet{idx}.xml",
                sheet_xml(sheet["headers"], rows, sheet["subtitle"]),
            )

    return OUTPUT_PATH


def add_rows(ws, rows: Iterable[dict]):
    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                row.get("module", ""),
                row.get("screen", ""),
                row.get("group", ""),
                row.get("label", ""),
                row.get("code", ""),
                row.get("description", ""),
                row.get("data_type", ""),
                row.get("required", ""),
                row.get("input_type", ""),
                row.get("decision", "Chưa xác nhận"),
                row.get("source", ""),
                row.get("dropdown_code", ""),
                row.get("formula", ""),
                row.get("depends_on", ""),
                row.get("notes", ""),
            ]
        )


lsx_fields = [
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Thông tin gốc", "label": "Mã LSX", "code": "code", "description": "Mã định danh lệnh sản xuất.", "data_type": "text", "required": "Có", "input_type": "Nhập tay + tự sinh gợi ý", "source": "Webapp hiện tại / production_headers", "formula": "Gợi ý theo ngày: DHAG-YYMMDD", "notes": "User cần xác nhận quy tắc mã chuẩn."},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Thông tin gốc", "label": "Mã hàng", "code": "sku", "description": "Mã sản phẩm/chủng loại sản xuất.", "data_type": "text", "required": "Có", "input_type": "Nhập tay", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Thông tin gốc", "label": "Tên hàng / diễn giải", "code": "productName", "description": "Tên sản phẩm hoặc mô tả nhận diện.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Thông tin gốc", "label": "Nơi nhận", "code": "destination", "description": "Bộ phận nhận lệnh / nơi dùng chứng từ.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_destinations"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Thông tin gốc", "label": "Số lượng", "code": "qtyPiece", "description": "Số viên/sợi/sản phẩm theo đơn hàng.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Chứng từ gốc", "label": "Ngày nghiệp vụ", "code": "occurredDate", "description": "Ngày phát sinh dự kiến của LSX.", "data_type": "date", "required": "Có", "input_type": "Date picker", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Chứng từ gốc", "label": "Số CT xuất", "code": "documentNo", "description": "Số chứng từ xuất gốc.", "data_type": "text", "required": "Không", "input_type": "Nhập tay / tự sinh", "source": "Webapp hiện tại / production_headers", "formula": "Nếu bỏ trống có thể tự sinh theo ngày + STT."},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Chứng từ gốc", "label": "Số CT nhập", "code": "documentInNo", "description": "Số chứng từ nhập gốc nếu có.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Chứng từ gốc", "label": "STT dòng", "code": "documentLineNo", "description": "Số thứ tự dòng trên chứng từ.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Chứng từ gốc", "label": "Loại giao dịch", "code": "movementType", "description": "Loại nghiệp vụ gốc của LSX.", "data_type": "enum", "required": "Có", "input_type": "Dropdown", "source": "Webapp hiện tại", "dropdown_code": "movement_type_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "Ngày kế hoạch", "code": "plannedDate", "description": "Ngày dự kiến bắt đầu/ghi nhận lệnh.", "data_type": "date", "required": "Không", "input_type": "Date picker", "source": "Webapp hiện tại / production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "Công đoạn dự kiến", "code": "plannedStage", "description": "Công đoạn chính dự kiến của lệnh.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_stages"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "Thợ / tổ dự kiến", "code": "plannedWorker", "description": "Người hoặc tổ được phân công.", "data_type": "text", "required": "Không", "input_type": "Dropdown động", "source": "workers", "dropdown_code": "workers_master"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "NVL dự kiến", "code": "plannedMaterial", "description": "Nguyên vật liệu chính dự kiến.", "data_type": "text", "required": "Không", "input_type": "Dropdown động", "source": "materials", "dropdown_code": "materials_master"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "Tuổi vàng dự kiến", "code": "plannedGoldAge", "description": "Tuổi vàng / purity dự kiến.", "data_type": "number", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "gold_age_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "Kế hoạch sản xuất", "label": "Loại nguyên liệu dự kiến", "code": "plannedMaterialType", "description": "Phân loại NVL/BTP/bột/phụ kiện.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "material_type_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Xuất gram", "code": "issued", "description": "Khối lượng xuất mặc định.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Nhập gram", "code": "returned", "description": "Khối lượng nhập mặc định.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Bột/phụ gram", "code": "powder", "description": "Khối lượng bột/phụ mặc định.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Chuyển gram", "code": "transferred", "description": "Khối lượng chuyển mặc định.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Tháng tính hao", "code": "lossPeriod", "description": "Kỳ dùng quyết toán hao hụt.", "data_type": "month", "required": "Có", "input_type": "Month picker", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Tháng NXT", "code": "nxtPeriod", "description": "Kỳ dùng cho báo cáo nhập xuất tồn.", "data_type": "month", "required": "Có", "input_type": "Month picker", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Mã nối NXT", "code": "nxtLinkCode", "description": "Mã nối chuỗi NXT hoặc nguồn vật tư.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_material_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Diễn giải giao dịch", "code": "sourceMaterialName", "description": "Mẫu diễn giải nghiệp vụ.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_movement_reasons"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Nguồn nhận NVL", "code": "sourceName", "description": "Loại kim loại/nguồn nhận NVL.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "material_metal_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Nguồn nhập", "code": "importSource", "description": "Nguồn nhập NVL.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Nguồn xuất", "code": "exportSource", "description": "Nguồn xuất NVL.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_options"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "TL quy KCP xuất", "code": "convertedIssueWeight", "description": "Trọng lượng xuất đã quy đổi.", "data_type": "number", "required": "Không", "input_type": "Công thức / cho phép sửa", "source": "production_headers", "formula": "issued * plannedGoldAge"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "TL quy KCP nhập", "code": "convertedReturnWeight", "description": "Trọng lượng nhập đã quy đổi.", "data_type": "number", "required": "Không", "input_type": "Công thức / cho phép sửa", "source": "production_headers", "formula": "returned * plannedGoldAge"},
    {"module": "Lệnh sản xuất", "screen": "Form tạo/cập nhật LSX", "group": "NXT / Tính hao", "label": "Ghi chú LSX", "code": "note", "description": "Lưu ý kỹ thuật hoặc yêu cầu sản xuất.", "data_type": "text", "required": "Không", "input_type": "Textarea", "source": "production_headers"},
    {"module": "Lệnh sản xuất", "screen": "Chi tiết LSX", "group": "Chỉ đọc / tổng hợp", "label": "Trạng thái LSX", "code": "status", "description": "Trạng thái tổng hợp từ các giao dịch NVL.", "data_type": "enum", "required": "Có", "input_type": "Chỉ đọc / hệ thống cập nhật", "source": "production_headers + material_movements", "formula": "Tổng hợp từ trạng thái giao dịch."},
]


nknvl_fields = [
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Tự sinh", "label": "Số CT tự sinh", "code": "documentNo", "description": "Số chứng từ tự sinh khi lưu nếu user không nhập.", "data_type": "text", "required": "Có", "input_type": "Tự sinh", "source": "Webapp hiện tại", "formula": "Theo ngày + STT lũy tiến trong ngày"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin LSX", "label": "Mã LSX", "code": "code", "description": "Lệnh sản xuất liên kết.", "data_type": "text", "required": "Có", "input_type": "Dropdown/nhập tay", "source": "production_orders"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin LSX", "label": "Mã hàng", "code": "sku", "description": "Mã sản phẩm trong giao dịch.", "data_type": "text", "required": "Có", "input_type": "Nhập tay / kế thừa LSX", "source": "production_orders"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin LSX", "label": "Tên hàng / diễn giải", "code": "productName", "description": "Tên hàng hoặc diễn giải nhận diện.", "data_type": "text", "required": "Không", "input_type": "Nhập tay / kế thừa LSX", "source": "production_orders"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin chứng từ", "label": "Ngày nghiệp vụ", "code": "occurredDate", "description": "Ngày phát sinh xuất/nhập NVL.", "data_type": "date", "required": "Có", "input_type": "Date picker", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin chứng từ", "label": "Nơi nhận", "code": "destination", "description": "Nơi nhận hoặc nguồn ghi nhận chứng từ.", "data_type": "text", "required": "Có", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_destinations"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin chứng từ", "label": "Số CT xuất", "code": "documentNo", "description": "Số chứng từ xuất.", "data_type": "text", "required": "Không", "input_type": "Nhập tay / tự sinh", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin chứng từ", "label": "Số CT nhập", "code": "documentInNo", "description": "Số chứng từ nhập nếu có.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Thông tin chứng từ", "label": "STT dòng", "code": "documentLineNo", "description": "STT dòng trong chứng từ.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NVL & công đoạn", "label": "Nguyên vật liệu", "code": "material", "description": "Tên NVL sử dụng trong giao dịch.", "data_type": "text", "required": "Có", "input_type": "Dropdown động", "source": "materials", "dropdown_code": "materials_master"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NVL & công đoạn", "label": "Công đoạn", "code": "stage", "description": "Mã công đoạn sản xuất.", "data_type": "text", "required": "Có", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_stages"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NVL & công đoạn", "label": "Thợ phụ trách", "code": "worker", "description": "Người nhận / phụ trách giao dịch.", "data_type": "text", "required": "Có", "input_type": "Dropdown động", "source": "workers", "dropdown_code": "workers_master"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Số lượng viên/sợi", "code": "qtyPiece", "description": "Số lượng thành phẩm/bán thành phẩm.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Xuất gram", "code": "issued", "description": "KCP xuất cho thợ.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Nhập gram", "code": "returned", "description": "Nhập về KCP.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Chuyển gram", "code": "transferred", "description": "Khối lượng chuyển tiếp.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "material_movements"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Loại giao dịch", "code": "movementType", "description": "Loại nghiệp vụ xuất/nhập/chuyển.", "data_type": "enum", "required": "Có", "input_type": "Dropdown", "source": "Webapp hiện tại", "dropdown_code": "movement_type_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Trạng thái tính hao", "code": "status", "description": "Tình trạng tính hao của giao dịch.", "data_type": "enum", "required": "Có", "input_type": "Dropdown", "source": "Webapp hiện tại", "dropdown_code": "status_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "Trọng lượng & trạng thái", "label": "Diễn giải giao dịch", "code": "sourceMaterialName", "description": "Mẫu diễn giải nghiệp vụ.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "journal_movement_reasons"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Tháng tính hao", "code": "lossPeriod", "description": "Kỳ quyết toán hao hụt.", "data_type": "month", "required": "Có", "input_type": "Month picker", "source": "material_movements", "formula": "carry-over theo ngày nghiệp vụ và trạng thái"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Tháng NXT", "code": "nxtPeriod", "description": "Kỳ nhập xuất tồn.", "data_type": "month", "required": "Có", "input_type": "Month picker", "source": "material_movements", "formula": "month(occurredDate)"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Tuổi vàng", "code": "goldAge", "description": "Purity của NVL.", "data_type": "number", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "gold_age_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Mã nối NXT", "code": "nxtLinkCode", "description": "Mã nối chuỗi NXT.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_material_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Loại nguyên liệu", "code": "materialType", "description": "Nhóm nguyên liệu/BTP/bột/phụ kiện.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "material_type_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Nguồn nhận NVL", "code": "sourceName", "description": "Nguồn kim loại / loại nguồn nhận.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "material_metal_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Nguồn nhập", "code": "importSource", "description": "Nguồn nhập NVL.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Nguồn xuất", "code": "exportSource", "description": "Nguồn xuất NVL.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "production-journal-options", "dropdown_code": "source_options"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "TL quy KCP xuất", "code": "convertedIssueWeight", "description": "Khối lượng xuất quy đổi.", "data_type": "number", "required": "Không", "input_type": "Công thức / cho phép sửa", "source": "material_movements", "formula": "issued * goldAge"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "TL quy KCP nhập", "code": "convertedReturnWeight", "description": "Khối lượng nhập quy đổi.", "data_type": "number", "required": "Không", "input_type": "Công thức / cho phép sửa", "source": "material_movements", "formula": "returned * goldAge"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Hao hụt tự tính", "code": "loss", "description": "Hao hụt phát sinh từ giao dịch.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "max(0, issued - returned - transferred)"},
    {"module": "Nhật ký NVL", "screen": "Form thêm giao dịch", "group": "NXT / Tính hao", "label": "Quy đổi 24K", "code": "loss_24k_preview", "description": "Hao hụt quy đổi để so sánh quyết toán.", "data_type": "number", "required": "Không", "input_type": "Công thức chỉ đọc", "source": "UI preview", "formula": "convertedIssueWeight - convertedReturnWeight"},
]


worker_box_fields = [
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Kỳ báo cáo", "code": "periodCode", "description": "Kỳ cần đối soát tồn thợ.", "data_type": "text", "required": "Có", "input_type": "Dropdown", "source": "workerBoxPeriods", "dropdown_code": "worker_box_periods"},
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Từ khóa tìm kiếm", "code": "query", "description": "Tìm theo thợ, công đoạn, NVL.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Trạng thái soát xét", "code": "reviewStatus", "description": "Lọc matched / pending / risk.", "data_type": "enum", "required": "Không", "input_type": "Dropdown", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_review_status"},
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Nhóm dòng", "code": "debtStatus", "description": "Treo nợ / dòng thường / đã xử lý.", "data_type": "enum", "required": "Không", "input_type": "Dropdown", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_debt_status"},
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Kim loại", "code": "metalCode", "description": "Lọc AU / AG / PT.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_metal_codes"},
    {"module": "Tồn hộp thợ", "screen": "Bộ lọc & đối soát", "group": "Bộ lọc", "label": "Tuổi vàng", "code": "goldAgeCode", "description": "Lọc theo tuổi vàng / nhóm purity.", "data_type": "text", "required": "Không", "input_type": "Dropdown", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_gold_age_codes"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Mã thợ", "code": "workerCode", "description": "Mã định danh người giữ tồn.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc/tính từ dữ liệu", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Tên thợ", "code": "workerName", "description": "Tên người giữ tồn.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc/tính từ dữ liệu", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Mã công đoạn", "code": "stageCode", "description": "Mã công đoạn của dòng tồn.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Tên công đoạn", "code": "stageName", "description": "Tên công đoạn sản xuất.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "NVL", "code": "materialName", "description": "Nguyên vật liệu đang đối soát.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Kim loại", "code": "metalCode", "description": "Nhóm kim loại AU/AG/PT.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Bảng tổng hợp", "group": "Chính", "label": "Tuổi vàng", "code": "goldAgeCode", "description": "Mã purity/tuổi vàng.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Tồn đầu kỳ", "code": "openingConvertedGram", "description": "Tồn đầu kỳ quy đổi.", "data_type": "number", "required": "Có", "input_type": "Công thức / snapshot", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Nhập trong kỳ", "code": "importConvertedGram", "description": "Tổng nhập trong kỳ quy đổi.", "data_type": "number", "required": "Có", "input_type": "Công thức / snapshot", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Xuất trong kỳ", "code": "exportConvertedGram", "description": "Tổng xuất trong kỳ quy đổi.", "data_type": "number", "required": "Có", "input_type": "Công thức / snapshot", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Tồn bột máy", "code": "machinePowderConvertedGram", "description": "Khối lượng bột/máy còn treo.", "data_type": "number", "required": "Không", "input_type": "Nhập/tính theo nghiệp vụ", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Tồn sổ sách", "code": "bookClosingConvertedGram", "description": "Tồn hệ thống sau tính toán.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "worker_box_balance_lines", "formula": "Tồn đầu + nhập - xuất +/- điều chỉnh"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Tồn thực tế", "code": "physicalConvertedGram", "description": "Tồn thực tế kiểm đếm.", "data_type": "number", "required": "Không", "input_type": "Nhập số", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Số liệu", "label": "Chênh lệch", "code": "diffConvertedGram", "description": "Sai lệch giữa sổ sách và thực tế.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "worker_box_balance_lines", "formula": "physicalConvertedGram - bookClosingConvertedGram"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Soát xét", "label": "Trạng thái soát xét", "code": "reviewStatus", "description": "Matched / pending / risk.", "data_type": "enum", "required": "Có", "input_type": "Dropdown/logic", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_review_status"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Soát xét", "label": "Trạng thái công nợ", "code": "debtStatus", "description": "treo_no / none / resolved.", "data_type": "enum", "required": "Không", "input_type": "Dropdown/logic", "source": "worker_box_balance_lines", "dropdown_code": "worker_box_debt_status"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Soát xét", "label": "XDC status", "code": "xdcStatus", "description": "Xác nhận đối chiếu XDC.", "data_type": "text", "required": "Không", "input_type": "Dropdown hoặc text", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Soát xét", "label": "NDC status", "code": "ndcStatus", "description": "Xác nhận đối chiếu NDC.", "data_type": "text", "required": "Không", "input_type": "Dropdown hoặc text", "source": "worker_box_balance_lines"},
    {"module": "Tồn hộp thợ", "screen": "Chi tiết dòng tồn", "group": "Soát xét", "label": "Nhận xét", "code": "comment", "description": "Ghi chú soát xét.", "data_type": "text", "required": "Không", "input_type": "Textarea", "source": "worker_box_balance_lines"},
]


loss_report_fields = [
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Bộ lọc / chọn", "label": "Công đoạn", "code": "stage", "description": "Trục tổng hợp chính của báo cáo hao hụt.", "data_type": "text", "required": "Có", "input_type": "Chọn dòng / dropdown tương lai", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Dòng phát sinh", "code": "count", "description": "Số giao dịch phát sinh trong công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "count(rows by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Loại vàng/NVL", "code": "material_count", "description": "Số loại vật tư phát sinh hao hụt trong công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "count(distinct material by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Tổng xuất", "code": "issued", "description": "Tổng khối lượng xuất theo công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "sum(issued by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Tổng nhập", "code": "returned", "description": "Tổng khối lượng nhập theo công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "sum(returned by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Hao hụt", "code": "loss", "description": "Tổng hao hụt thực tế theo công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "sum(loss by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Tổng hợp theo công đoạn", "group": "Tổng hợp", "label": "Hao hụt quy 24K", "code": "convertedLoss", "description": "Hao hụt quy đổi phục vụ quyết toán.", "data_type": "number", "required": "Có", "input_type": "Công thức chỉ đọc", "source": "material_movements", "formula": "sum(loss * goldAge by stage)"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết công đoạn", "group": "Theo loại vàng/NVL", "label": "Loại vàng / NVL", "code": "materialKey", "description": "Loại vàng hoặc NVL gây hao hụt trong công đoạn.", "data_type": "text", "required": "Có", "input_type": "Công thức nhóm", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết công đoạn", "group": "Theo loại vàng/NVL", "label": "Số dòng", "code": "count", "description": "Số giao dịch phát sinh của nhóm NVL trong công đoạn.", "data_type": "number", "required": "Có", "input_type": "Công thức", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết công đoạn", "group": "Theo loại vàng/NVL", "label": "Số thợ", "code": "workers", "description": "Số thợ có phát sinh hao hụt trong nhóm.", "data_type": "number", "required": "Có", "input_type": "Công thức", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết công đoạn", "group": "Theo loại vàng/NVL", "label": "Số LSX", "code": "lsxCodes", "description": "Số LSX phát sinh trong nhóm.", "data_type": "number", "required": "Có", "input_type": "Công thức", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết công đoạn", "group": "Theo loại vàng/NVL", "label": "Bột", "code": "powder", "description": "Tổng bột/phụ theo nhóm.", "data_type": "number", "required": "Không", "input_type": "Công thức", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết giao dịch", "group": "Truy ngược", "label": "LSX", "code": "code", "description": "LSX gốc của giao dịch hao hụt.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết giao dịch", "group": "Truy ngược", "label": "Mã hàng", "code": "sku", "description": "Mã hàng của giao dịch.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết giao dịch", "group": "Truy ngược", "label": "NVL / Loại vàng", "code": "material", "description": "NVL thực tế phát sinh hao hụt.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết giao dịch", "group": "Truy ngược", "label": "Thợ", "code": "worker", "description": "Thợ phát sinh giao dịch.", "data_type": "text", "required": "Có", "input_type": "Chỉ đọc", "source": "material_movements"},
    {"module": "Báo cáo hao hụt", "screen": "Chi tiết giao dịch", "group": "Truy ngược", "label": "Trạng thái", "code": "status", "description": "Đang xử lý / Treo nợ / Xác định / Đã chốt.", "data_type": "enum", "required": "Có", "input_type": "Chỉ đọc", "source": "material_movements", "dropdown_code": "status_options"},
]


pricing_fields = [
    {"module": "Giá & định mức", "screen": "Bảng giá / định mức", "group": "Danh mục giá", "label": "Kim loại / NVL", "code": "metal", "description": "Đối tượng duyệt giá.", "data_type": "text", "required": "Có", "input_type": "Dropdown/master", "source": "priceRows/materials"},
    {"module": "Giá & định mức", "screen": "Bảng giá / định mức", "group": "Danh mục giá", "label": "Nguồn giá", "code": "source", "description": "Nguồn hình thành giá.", "data_type": "text", "required": "Không", "input_type": "Nhập tay", "source": "priceRows"},
    {"module": "Giá & định mức", "screen": "Bảng giá / định mức", "group": "Danh mục giá", "label": "Purity", "code": "purity", "description": "Tuổi vàng / độ tinh khiết áp dụng.", "data_type": "number/text", "required": "Không", "input_type": "Dropdown/nhập", "source": "priceRows"},
    {"module": "Giá & định mức", "screen": "Bảng giá / định mức", "group": "Danh mục giá", "label": "Giá duyệt", "code": "value", "description": "Giá đã duyệt theo kỳ.", "data_type": "currency", "required": "Có", "input_type": "Nhập số", "source": "priceRows"},
    {"module": "Giá & định mức", "screen": "Bảng giá / định mức", "group": "Danh mục giá", "label": "Trạng thái duyệt", "code": "status", "description": "Đã duyệt / Chờ duyệt / Nháp.", "data_type": "enum", "required": "Có", "input_type": "Dropdown", "source": "priceRows"},
    {"module": "Giá & định mức", "screen": "Danh mục NVL", "group": "Master", "label": "Mã NVL", "code": "code", "description": "Mã danh mục nguyên vật liệu.", "data_type": "text", "required": "Có", "input_type": "Nhập tay", "source": "materials"},
    {"module": "Giá & định mức", "screen": "Danh mục NVL", "group": "Master", "label": "Tên NVL", "code": "name", "description": "Tên hiển thị của nguyên vật liệu.", "data_type": "text", "required": "Có", "input_type": "Nhập tay", "source": "materials"},
    {"module": "Giá & định mức", "screen": "Danh mục NVL", "group": "Master", "label": "Nhóm", "code": "category", "description": "gold / silver / platinum / other.", "data_type": "text", "required": "Có", "input_type": "Dropdown/nhập", "source": "materials"},
    {"module": "Giá & định mức", "screen": "Danh mục NVL", "group": "Master", "label": "Purity", "code": "purity", "description": "Hàm lượng quy đổi.", "data_type": "number", "required": "Có", "input_type": "Nhập số", "source": "materials"},
    {"module": "Giá & định mức", "screen": "Danh mục NVL", "group": "Master", "label": "Đơn vị", "code": "unit", "description": "gram / chỉ / lượng.", "data_type": "text", "required": "Có", "input_type": "Nhập tay/dropdown", "source": "materials"},
]


dropdown_rows = []


def add_dropdown(group: str, module: str, field_code: str, options: list[tuple[str, str, str]]):
    for idx, (value, label, note) in enumerate(options, start=1):
        dropdown_rows.append({
            "group": group,
            "module": module,
            "field_code": field_code,
            "order": idx,
            "value": value,
            "label": label,
            "note": note,
        })


add_dropdown("journal_destinations", "LSX / Nhật ký NVL", "destination", [
    ("NKBC", "NKBC - Nhật ký bạc/công đoạn", ""),
    ("BKNXT", "BKNXT - Bảng kê NXT", ""),
    ("KCP", "KCP - Kho cấp phát", ""),
])
add_dropdown("journal_stages", "LSX / Nhật ký NVL / Báo cáo hao hụt", "stage", [
    ("CKE", "CKE - Cán kéo", ""),
    ("CDT", "CDT - Cán dát", ""),
    ("DAN", "DAN - Đan dây", ""),
    ("BIEN", "BIEN - Biến", ""),
    ("QBI", "QBI - Quay bi", ""),
    ("BAO", "BAO - Bào dây", ""),
    ("PI", "PI - Pi", ""),
    ("DAP", "DAP - Dập định hình", ""),
    ("DKB", "DKB - Đánh bóng", ""),
    ("GEP", "GEP - Ghép dây", ""),
    ("NAU", "NAU - Nấu", ""),
    ("SXK", "SXK - Sản xuất khóa", ""),
    ("HTH", "HTH - Hoàn thiện", ""),
])
add_dropdown("movement_type_options", "LSX / Nhật ký NVL", "movementType", [
    ("issue", "Xuất", ""),
    ("return", "Nhập", ""),
    ("transfer", "Chuyển", ""),
    ("adjustment", "Điều chỉnh", ""),
])
add_dropdown("status_options", "Nhật ký NVL / Báo cáo hao hụt", "status", [
    ("Đang xử lý", "Đang xử lý", ""),
    ("Treo nợ", "Treo nợ", ""),
    ("Xác định", "Xác định", "Bắt buộc cho các công đoạn trực tiếp tính công nợ theo rule hiện tại."),
    ("Đã chốt", "Đã chốt", ""),
])
add_dropdown("gold_age_options", "LSX / Nhật ký NVL", "goldAge", [
    ("0.9999", "24K / 99.99%", ""),
    ("0.9583", "23K / 95.83%", ""),
    ("0.75", "18K / 75.00%", ""),
    ("0.61", "15K / 61.00%", ""),
    ("0.925", "Bạc 92.5%", ""),
])
add_dropdown("source_material_options", "LSX / Nhật ký NVL", "nxtLinkCode", [
    ("750Y", "750Y - Vàng 18K vàng", ""),
    ("750W", "750W - Vàng 18K trắng", ""),
    ("750R", "750R - Vàng 18K hồng", ""),
    ("BAC925", "BAC925 - Bạc 92.5", ""),
    ("Nấu mới", "Nấu mới", ""),
    ("Nấu quay đầu", "Nấu quay đầu", ""),
    ("TP", "TP - Thành phẩm/BTP nhập lại", ""),
])
add_dropdown("material_type_options", "LSX / Nhật ký NVL", "materialType", [
    ("NL18K", "NL18K - Nguyên liệu vàng 18K", ""),
    ("NL24K", "NL24K - Nguyên liệu vàng 24K", ""),
    ("NL23K", "NL23K - Nguyên liệu vàng 23K", ""),
    ("NLBAC92.5", "NLBAC92.5 - Nguyên liệu bạc 92.5", ""),
    ("NLBAC9999", "NLBAC9999 - Bạc 9999", ""),
    ("BOT18K", "BOT18K - Bột vàng 18K", ""),
    ("BOT23K", "BOT23K - Bột vàng 23K", ""),
    ("BOT24K", "BOT24K - Bột vàng 24K", ""),
    ("PK18K", "PK18K - Phụ kiện 18K", ""),
    ("BTPDAY18K", "BTPDAY18K - BTP dây 18K", ""),
    ("BTPDAYBAC92.5", "BTPDAYBAC92.5 - BTP dây bạc", ""),
])
add_dropdown("material_metal_options", "LSX / Nhật ký NVL / Tồn hộp thợ", "sourceName", [
    ("AU", "AU - Vàng", ""),
    ("AG", "AG - Bạc", ""),
    ("PT", "PT - Platinum", ""),
])
add_dropdown("source_options", "LSX / Nhật ký NVL", "importSource/exportSource", [
    ("CĐ", "CĐ - Công đoạn", ""),
    ("VN", "VN - Nguồn Việt Nam", ""),
    ("PHAN_KIM", "Phân kim", ""),
    ("US", "Nhập từ US", ""),
    ("PL", "PL - Phân loại/phiếu lẻ", ""),
    ("KCP", "KCP - Kho cấp phát", ""),
])
add_dropdown("journal_movement_reasons", "LSX / Nhật ký NVL", "sourceMaterialName", [
    ("Xuất cán kéo", "Xuất cán kéo", ""),
    ("Nhập sau cán kéo", "Nhập sau cán kéo", ""),
    ("Xuất đan dây", "Xuất đan dây", ""),
    ("Nhập sau đan dây", "Nhập sau đan dây", ""),
    ("Xuất quay bi", "Xuất quay bi", ""),
    ("Nhập sau quay bi", "Nhập sau quay bi", ""),
    ("Xuất bào dây", "Xuất bào dây", ""),
    ("Nhập sau bào dây", "Nhập sau bào dây", ""),
    ("Xuất nấu mới", "Xuất nấu mới", ""),
    ("Nhập sau nấu mới", "Nhập sau nấu mới", ""),
    ("Xuất nấu lại", "Xuất nấu lại", ""),
    ("Nhập sau nấu lại", "Nhập sau nấu lại", ""),
])
add_dropdown("worker_box_review_status", "Tồn hộp thợ", "reviewStatus", [
    ("matched", "Đã khớp", ""),
    ("pending", "Cần soát", ""),
    ("risk", "Rủi ro", ""),
])
add_dropdown("worker_box_debt_status", "Tồn hộp thợ", "debtStatus", [
    ("treo_no", "Treo nợ", ""),
    ("none", "Dòng thường", ""),
    ("resolved", "Đã xử lý", ""),
])


logic_rows = [
    ["Mã LSX gợi ý", "Lệnh sản xuất", "buildProductionOrderCode('DHAG', occurredDate)", "Sinh mã LSX theo ngày khi tạo mới."],
    ["Số CT tự sinh", "Nhật ký NVL", "YYYYMMDD + STT lũy tiến trong ngày", "Chỉ tự sinh khi user không nhập tay."],
    ["Hao hụt giao dịch", "Nhật ký NVL", "max(0, issued - returned - transferred)", "Công thức hiện có trong webapp."],
    ["TL quy KCP xuất", "LSX / Nhật ký NVL", "issued * goldAge", "Quy đổi theo tuổi vàng/purity."],
    ["TL quy KCP nhập", "LSX / Nhật ký NVL", "returned * goldAge", "Quy đổi theo tuổi vàng/purity."],
    ["Quy đổi hao hụt 24K", "Báo cáo hao hụt", "loss * purity", "Hiện dùng purity của giao dịch để quy về chuẩn so sánh."],
    ["Trạng thái LSX tổng hợp", "Lệnh sản xuất", "Nếu có Treo nợ => Treo nợ; nếu có Đang xử lý => Đang xử lý; nếu tất cả Đã chốt => Đã chốt; còn lại => Xác định", "Logic tổng hợp từ status giao dịch."],
    ["Tháng NXT", "LSX / Nhật ký NVL", "month(occurredDate)", "Mặc định theo ngày phát sinh."],
    ["Tháng tính hao", "LSX / Nhật ký NVL", "getCarryOverLossPeriod(occurredDate, status)", "Có carry-over cho case gối đầu."],
    ["Tồn sổ sách", "Tồn hộp thợ", "Tồn đầu + nhập - xuất +/- điều chỉnh", "Hiện là snapshot tổng hợp theo kỳ."],
    ["Chênh lệch tồn", "Tồn hộp thợ", "physicalConvertedGram - bookClosingConvertedGram", "Dùng để phân loại matched/pending/risk."],
]


review_rows = [
    ["Lệnh sản xuất", "Mã LSX", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Lệnh sản xuất", "Mã hàng", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Lệnh sản xuất", "Nơi nhận", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Nhật ký NVL", "Công đoạn", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Nhật ký NVL", "Thợ phụ trách", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Nhật ký NVL", "Loại giao dịch", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Nhật ký NVL", "Trạng thái tính hao", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Tồn hộp thợ", "Tồn thực tế", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Báo cáo hao hụt", "Nhóm theo công đoạn", "", "Giữ/Bỏ/Sửa logic", "", ""],
    ["Báo cáo hao hụt", "Nhóm theo loại vàng/NVL", "", "Giữ/Bỏ/Sửa logic", "", ""],
]


def build_sheet(ws, title: str, subtitle: str, rows: list[dict]):
    ws.title = title
    ws.append([subtitle])
    ws.append(FIELD_HEADERS)
    for cell in ws[1]:
        cell.font = bold_font
        cell.fill = section_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(FIELD_HEADERS))
    style_header(ws, row=2)
    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                row.get("module", ""),
                row.get("screen", ""),
                row.get("group", ""),
                row.get("label", ""),
                row.get("code", ""),
                row.get("description", ""),
                row.get("data_type", ""),
                row.get("required", ""),
                row.get("input_type", ""),
                row.get("decision", "Chưa xác nhận"),
                row.get("source", ""),
                row.get("dropdown_code", ""),
                row.get("formula", ""),
                row.get("depends_on", ""),
                row.get("notes", ""),
            ]
        )
    autosize(ws)
    ws.freeze_panes = "A3"


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "output_xlsx": str(OUTPUT_PATH),
        "sheets": [
            {"name": "01_LSX_Fields", "subtitle": "Field mapping cho module Lệnh sản xuất", "headers": FIELD_HEADERS, "rows": lsx_fields},
            {"name": "02_NKNVL_Fields", "subtitle": "Field mapping cho module Nhật ký NVL", "headers": FIELD_HEADERS, "rows": nknvl_fields},
            {"name": "03_TonHopTho", "subtitle": "Field mapping cho module Tồn hộp thợ", "headers": FIELD_HEADERS, "rows": worker_box_fields},
            {"name": "04_HaoHut", "subtitle": "Field mapping cho module Báo cáo hao hụt", "headers": FIELD_HEADERS, "rows": loss_report_fields},
            {"name": "05_GiaDinhMuc", "subtitle": "Field mapping cho module Giá & định mức", "headers": FIELD_HEADERS, "rows": pricing_fields},
            {
                "name": "06_DropdownLists",
                "subtitle": "Danh sách dropdown hiện tại để user xác nhận lại trước khi triển khai",
                "headers": ["STT", "Dropdown code", "Module", "Field áp dụng", "Value", "Label", "Ghi chú"],
                "rows": dropdown_rows,
            },
            {
                "name": "07_Logic_CongThuc",
                "subtitle": "Các công thức và logic hệ thống cần user xác nhận",
                "headers": ["STT", "Tên logic", "Module", "Công thức / Rule", "Mô tả"],
                "rows": logic_rows,
            },
            {
                "name": "08_Review_User",
                "subtitle": "Sheet để user quyết định giữ/bỏ/sửa từng nhóm field trước khi triển khai",
                "headers": ["STT", "Module", "Field / nhóm xác nhận", "Quyết định", "Người xác nhận", "Ngày", "Ghi chú"],
                "rows": review_rows,
            },
        ],
    }
    JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    if Workbook is None:
        build_xlsx_from_payload(payload)
        print(OUTPUT_PATH)
        return

    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb.create_sheet(), "01_LSX_Fields", "Field mapping cho module Lệnh sản xuất", lsx_fields)
    build_sheet(wb.create_sheet(), "02_NKNVL_Fields", "Field mapping cho module Nhật ký NVL", nknvl_fields)
    build_sheet(wb.create_sheet(), "03_TonHopTho", "Field mapping cho module Tồn hộp thợ", worker_box_fields)
    build_sheet(wb.create_sheet(), "04_HaoHut", "Field mapping cho module Báo cáo hao hụt", loss_report_fields)
    build_sheet(wb.create_sheet(), "05_GiaDinhMuc", "Field mapping cho module Giá & định mức", pricing_fields)

    ws_dropdown = wb.create_sheet("06_DropdownLists")
    ws_dropdown.append(["Danh sách dropdown hiện tại để user xác nhận lại trước khi triển khai"])
    ws_dropdown.merge_cells("A1:G1")
    for cell in ws_dropdown[1]:
        cell.font = bold_font
        cell.fill = section_fill
    ws_dropdown.append(["STT", "Dropdown code", "Module", "Field áp dụng", "Value", "Label", "Ghi chú"])
    style_header(ws_dropdown, row=2)
    for idx, row in enumerate(dropdown_rows, start=1):
        ws_dropdown.append([idx, row["group"], row["module"], row["field_code"], row["value"], row["label"], row["note"]])
    autosize(ws_dropdown)
    ws_dropdown.freeze_panes = "A3"

    ws_logic = wb.create_sheet("07_Logic_CongThuc")
    ws_logic.append(["Các công thức và logic hệ thống cần user xác nhận"])
    ws_logic.merge_cells("A1:D1")
    for cell in ws_logic[1]:
        cell.font = bold_font
        cell.fill = section_fill
    ws_logic.append(["STT", "Tên logic", "Module", "Công thức / Rule", "Mô tả"])
    style_header(ws_logic, row=2)
    for idx, row in enumerate(logic_rows, start=1):
        ws_logic.append([idx, row[0], row[1], row[2], row[3]])
    autosize(ws_logic)
    ws_logic.freeze_panes = "A3"

    ws_review = wb.create_sheet("08_Review_User")
    ws_review.append(["Sheet để user quyết định giữ/bỏ/sửa từng nhóm field trước khi triển khai"])
    ws_review.merge_cells("A1:F1")
    for cell in ws_review[1]:
        cell.font = bold_font
        cell.fill = section_fill
    ws_review.append(["STT", "Module", "Field / nhóm xác nhận", "Quyết định", "Người xác nhận", "Ngày", "Ghi chú"])
    style_header(ws_review, row=2)
    for idx, row in enumerate(review_rows, start=1):
        ws_review.append([idx, row[0], row[1], row[3], row[4], row[5], row[2]])
    autosize(ws_review)
    ws_review.freeze_panes = "A3"

    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
